'''
1 - Entrar na planilha e extrair cpf do cliente
2 - Entro no site https://consultcpf-devaprender.netlify.app/ e uso o cpf da planilha para pesquisar o status do pagementeo e a data
3 - Verificar se esta "em dia" ou "atrasada"
4 - Se estiver "em dia", pegar a data do pagamento e o metodo de pagamento
5 - Caso contrario ( se estiver atrasado ), colocar status pendente
6 - Inserir essas informações (nome, valor, cpf, vencimento, status e caso esteja em dia , data do pagament, metodo do pagemnto colocar "ok", caso esteja atrasada "pendente")
7 - Inserir até chegar no ultimo cliente
'''
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep

# 1 - Entrar na planilha e extrair cpf do cliente
planilha_clientes = openpyxl.load_workbook('dados_clientes.xlsx')
pagina_clientes = planilha_clientes['Sheet1']

# 2 - Entrar no site https://consultcpf-devaprender.netlify.app/ e uso o cpf da planilha para pesquisar o status do pagementeo e a data
driver = webdriver.Chrome()
driver.get('https://consultcpf-devaprender.netlify.app/')

for linha in pagina_clientes.iter_rows(min_row=2, values_only=True):
    nome, valor, cpf, vencimento = linha

    sleep(5)
    campo_pesquisa = driver.find_element(By.XPATH, "//input[@id='cpfInput']")
    sleep(1)
    campo_pesquisa.clear()
    campo_pesquisa.send_keys(cpf)
    sleep(1)
    # 3 - Verificar se esta "em dia" ou "atrasada"
    botao_pesquisar = driver.find_element(
        By.XPATH, "//button[@class='btn btn-custom btn-lg btn-block mt-3']")
    sleep(1)
    botao_pesquisar.click()
    sleep(4)

    status = driver.find_element(By.XPATH, "//span[@id='statusLabel']")
    if status.text == 'em dia':
        # 4 - Se estiver "em dia", pegar a data do pagamento e o metodo de pagamento
        data_pagamento = driver.find_element(
            By.XPATH, "//p[@id='paymentDate']")
        metodo_pagamento = driver.find_element(
            By.XPATH, "//p[@id='paymentMethod']")

        data_payment = data_pagamento.text.split()[3]
        metodo_payment = metodo_pagamento.text.split()[3]

        planilha_clientes = openpyxl.load_workbook('dados_clientes.xlsx')
        pagina_fechamento = planilha_clientes['Sheet2']

        pagina_fechamento.append(
            [nome, valor, cpf, vencimento, 'Em dia', data_payment, metodo_payment])

        planilha_clientes.save('dados_clientes.xlsx')

    else:
        # 5 - Caso contrario ( se estiver atrasado ), colocar status pendente
        planilha_clientes = openpyxl.load_workbook('dados_clientes.xlsx')
        pagina_fechamento = planilha_clientes['Sheet2']

        pagina_fechamento.append([nome, valor, cpf, vencimento, 'Pendente'])
        planilha_clientes.save('dados_clientes.xlsx')
